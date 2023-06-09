import cv2
from openpyxl import load_workbook, Workbook
import pytesseract
from PIL import Image, ImageEnhance
import re
import youtube_dl
import os
import uuid
from tqdm import tqdm

def ocr_function(image):
    # Image preprocessing for OCR
    pil_image = Image.fromarray(image)
    pil_image = pil_image.convert("RGB")
    enhancer = ImageEnhance.Contrast(pil_image)
    pil_image = enhancer.enhance(2.0)

    # OCR using Tesseract
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    custom_config = r'--oem 1 --psm 7 --tessdata-dir "C:\Users\Nikita\Desktop\ligapro.parser\training"'
    text = pytesseract.image_to_string(pil_image, config=custom_config, lang='eng')

    return text

def parse_score_changes(video_path, skip_frames, row_index):
    # Load video
    video = cv2.VideoCapture(video_path)

    # Define the score box region of interest (top left point and dimensions)
    score_box1_top_left = (80, 40)
    score_box1_width = 107
    score_box1_height = 40

    score_box2_top_left = (430, 40)
    score_box2_width = 72
    score_box2_height = 40

    # Initialize workbook and sheet
    try:
        wb = load_workbook(r"C:\Users\Nikita\Desktop\ligapro.parser\score_changes.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    # Create a new sheet for the current video
    video_index = len(wb.sheetnames) + 1
    sheet_name = f"Video {video_index}"
    sheet = wb.create_sheet(title=sheet_name)

    # Initialize column index
    col_index = 1

    # Variables to store previous score and time of last score change
    prev_score = ""
    prev_time = ""
    last_score_change_time = None
    zero_zero = False

    # Iterate over each frame of the video
    total_frames = int(video.get(cv2.CAP_PROP_FRAME_COUNT))
    with tqdm(total=total_frames, ncols=80, unit="frame") as pbar:
        frame_count = 0
        while True:
            ret, frame = video.read()
            if not ret:
                break
            frame_count += 1
            if frame_count % skip_frames != 0:
                continue
            current_time = video.get(cv2.CAP_PROP_POS_MSEC) / 1000

            # Extract the score box from the frame
            score_box1 = frame[
                         score_box1_top_left[1]: score_box1_top_left[1] + score_box1_height,
                         score_box1_top_left[0]: score_box1_top_left[0] + score_box1_width,
                         ]

            # Extract the second score box from the frame
            score_box2 = frame[
                         score_box2_top_left[1]: score_box2_top_left[1] + score_box2_height,
                         score_box2_top_left[0]: score_box2_top_left[0] + score_box2_width,
                         ]

            # Convert the score boxes to grayscale
            score_box1_gray = cv2.cvtColor(score_box1, cv2.COLOR_BGR2GRAY)
            score_box2_gray = cv2.cvtColor(score_box2, cv2.COLOR_BGR2GRAY)

            # Convert the grayscale images to binary (black and white) using thresholding
            _, score_box1_bw = cv2.threshold(score_box1_gray, 127, 255, cv2.THRESH_BINARY)
            _, score_box2_bw = cv2.threshold(score_box2_gray, 127, 255, cv2.THRESH_BINARY)

            # Perform OCR to extract the score time from the first score box
            time = ocr_function(score_box1_bw)[:3].replace(':', '')

            # Perform OCR to extract the score time from the second score box
            score = ocr_function(score_box2_bw).replace("\n", "").replace(" ", "")
            if score and score != prev_score:
                try:
                    if len(re.findall(r'\d', score)) != 2 or not ('-' in score):
                        continue
                    if (prev_score != "" and score != "") and (
                            not (score[0].isdigit())
                            or not (score[2].isdigit())
                            or len(re.findall(r'\d', time)) < 2
                            or not (time[0].isdigit())
                            or not (time[1].isdigit())
                            or (int(score[0]) + int(score[2])) - (int(prev_score[0]) + int(prev_score[2])) >= 2
                            or (int(prev_score[0]) + int(prev_score[2])) - (int(score[0]) + int(score[2])) >= 2
                    ):
                        continue
                except:
                    pass
                if score == "0-0":
                    if last_score_change_time is not None:
                        row_index += 1  # Move to the next row
                        col_index = 1  # Reset the column index
                else:
                    # Write the score change in a separate cell
                    sheet.cell(row=row_index, column=col_index).value = score + f" ({time})"
                    col_index += 1

                print(f"Score changed: {prev_score} ({time}) -> {score} ({time})")
                prev_time = time
                prev_score = score
                last_score_change_time = current_time

            pbar.update(skip_frames)  # Update the progress bar
            try:
                if int(time) < 90 and int(time) > 10:
                    zero_zero = True
            except:
                pass
            if zero_zero:
                if score and time.isdigit() and int(time) >= 90:
                    prev_score = score
                    prev_time = time
                if score and score == "0-0" and prev_score == "0-0" and time.isdigit() \
                    and int(prev_time) >= 90 and int(time[0]) == 0:
                    row_index += 1
                    zero_zero = False
    # Save the Excel file
    wb.save(r"C:\Users\Nikita\Desktop\ligapro.parser\score_changes.xlsx")
    print('Парсинг завершен и файл Excel сохранен успешно')

    # Release the video file
    video.release()

def download_and_parse_video(urls, skip_frames):
    video_urls = urls.split(',')

    # Initialize the row index
    row_index = 1

    # Iterate over each URL
    for url in video_urls:
        url = url.strip()

        if 'www' in url:
            try:
                video_folder = os.path.join('C:\\', 'Users', 'Nikita', 'Desktop', 'ligapro.parser', 'videos')
                video_hash = str(uuid.uuid4())
                video_filename = os.path.join(video_folder, f"video_{video_hash}.mp4")
                ydl_opts = {
                    'format': 'bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
                    'continuedl': True,
                    'outtmpl': video_filename,
                    'nocheckcertificate': True
                }
                with youtube_dl.YoutubeDL(ydl_opts) as ydl:
                    ydl.download([url])
            except youtube_dl.utils.DownloadError as e:
                print(f"Error occurred during video download: {str(e)}")
                break
        else:
            video_filename = url

        # Parse the video for score changes
        parse_score_changes(video_filename, skip_frames, row_index)

video_urls = input('Ожидаю URL для видео (браузерная или полная локальная ссылка): ').replace("\"", '')
skip_frames = int(input("Сколько кадров пропускать: "))
download_and_parse_video(video_urls, skip_frames)